import json
import logging
import ntpath
from typing import List

from openpyxl import load_workbook

from labelregions.LabelRegion import LabelRegion
from labelregions.LabelRegionType import LabelRegionType

logger = logging.getLogger(__name__)


class LabelRegionPreprocessor(object):
    def __init__(self, remove_empty_cells=True):
        self.remove_empty_cells = remove_empty_cells

    def _read_annotations(self):
        """Returns all annotations as dict"""
        with open(self._annotation_file) as f:
            data = f.read()
        return json.loads(data)

    def _get_annotation(self):
        """Returns the annotation of the spreadsheet"""
        annotations = self._read_annotations()
        spreadsheet_name = ntpath.basename(self._spreadsheet_file)
        # Only load first sheet, can not decide relevance of other sheets
        annotation_key = spreadsheet_name + '_' + self._sheet_name + '.csv'
        return annotations[annotation_key]

    def _flatten_and_rewrite_label_regions(self, annotation):
        """Flattens the LR element list and clears/replaces labels to
        fit the table model (Header, Data).
        Returns a dict mapping each labale region name to its data dict"""
        # Flatten annotated regions, as we dont want the table ground truth data
        label_regions = []
        for region in annotation["regions"]:
            label_regions.extend(region["elements"])

        # Filter / rename specific labels
        # Drop MetaTitles
        label_regions = [lr for lr in label_regions if lr["type"] != "MetaTitle"]
        # Set Derived = Data and GroupHead = Header
        for lr in label_regions:
            if lr["type"] == "Derived":
                lr["type"] = "Data"
            if lr["type"] == "GroupHead":
                lr["type"] = "Header"

        label_regions_dict = {}
        for lr in label_regions:
            label_regions_dict[lr["element_label"]] = lr
        return label_regions_dict

    def _split_lrs_into_cells(self, flat_lrs):
        """Returns an array of rows, each containing annotated cells, in order"""
        cell_rows = {}
        for region_name, region_data in flat_lrs.items():
            min_x, min_y = region_data['top_lx']
            max_x, max_y = region_data['bot_rx']
            for y in range(min_y, max_y + 1):
                row = []
                for x in range(min_x, max_x + 1):
                    cell = {
                        "type": region_data['type'],
                        "x": x,
                        "y": y,
                    }
                    cell_rows[y] = cell_rows.get(y, []) + [cell]
        # cell_rows now contains potentially unordered cells of a single row
        # order them
        rows = []
        for row_id in sorted(cell_rows.keys()):
            row = sorted(cell_rows[row_id], key=lambda single_cell: single_cell["x"])
            rows.append(row)

        return rows

    def _remove_empty_cells(self, cell_rows):
        """Removes empty cells from given rows
        The paper assumes that only non-empty cells are annotated
        In our model annotations can span empty cells"""
        filtered_cell_rows = []
        for row in cell_rows:
            filtered_row = []
            for cell in row:
                if not self._cell_empty(cell):
                    filtered_row.append(cell)
            filtered_cell_rows.append(filtered_row)

        return filtered_cell_rows

    def _cell_empty(self, cell):
        """Returns whether the value at this cell is
        empty (only whitespace)"""

        # Increase coordinates by one as our annotations are indexed
        # starting from 0 and openpyxl is indexted starting from 1
        return self._worksheet.cell(
            row=cell["y"] + 1,
            column=cell["x"] + 1,
        ).value is None

    def _merge_labled_cells_into_lrs(self, cell_rows) -> List[LabelRegion]:
        """Spreads cells to strictly rectangular regions
        Refer to https://upcommons.upc.edu/bitstream/handle/2117/128001/ROMERO%20Table%20recognition.pdf;jsessionid=F7D1099DA22A66950693F51EE0720A5C?sequence=1
        E. Koci, M. Thiele, W. Lehner and O. Romero, "Table Recognition in Spreadsheets via a Graph Representation," 2018 13th IAPR International Workshop on Document Analysis Systems (DAS), 2018, pp. 139-144, doi: 10.1109/DAS.2018.48.
        for alogrightm description
        """
        # Annotate cells whether they want to form a sequence with their right neighbour
        rows_with_merge_annotations = []
        for row in cell_rows:
            for i, cell in enumerate(row):
                if (i < len(row) - 1):
                    neighbours = cell['x'] + 1 == row[i + 1]['x']
                    same_type = cell['type'] == row[i + 1]['type']
                    if neighbours and same_type:
                        row[i + 1]['merge_left'] = True
            rows_with_merge_annotations.append(row)

        # Create a new sequence for a cell that does not want to be merged
        # Add all other cells to the current sequence (as cells are in order)
        rows_with_sequences = []
        for row in rows_with_merge_annotations:
            new_row = []
            for cell in row:
                if cell.get("merge_left", False) is False:
                    # Cell starts a new sequence
                    sequence = {
                        "type": cell["type"],
                        "start_x": cell["x"],
                        "stop_x": cell["x"],
                        "y": cell["y"],
                    }
                    new_row.append(sequence)
                else:
                    # Cell adds to last sequence
                    new_row[-1]["stop_x"] = cell["x"]
            rows_with_sequences.append(new_row)

        # Annotate horizontal sequences for merging
        next_lr_id = 0
        for y, row in enumerate(rows_with_sequences):
            for sequence in row:
                if sequence.get("lr_id", False) is False:
                    # Sequence does not belong to a lr yet
                    sequence["lr_id"] = next_lr_id
                    next_lr_id += 1
                # Look for sequences below that match out label and bounds
                if y < len(rows_with_sequences) - 1:
                    for sequence_below in rows_with_sequences[y + 1]:
                        same_type = sequence["type"] == sequence_below["type"]
                        same_min_x = sequence["start_x"] == sequence_below["start_x"]
                        same_max_x = sequence["stop_x"] == sequence_below["stop_x"]
                        if same_type and same_min_x and same_max_x:
                            # Align and same label, belong to same sequence
                            # We do not need to check if lr_id is already set
                            # because each sequence belongs to exactly one lr
                            sequence_below["lr_id"] = sequence["lr_id"]
                            # We can break after we found a match in the row below
                            # Because there can not be more matches per row
                            break

        # Flatten and group sequences
        lr_to_parts = {}
        for row in rows_with_sequences:
            for sequence in row:
                lr_to_parts[sequence["lr_id"]] = lr_to_parts.get(sequence["lr_id"], []) + [sequence]

        # 'Merge' lr parts
        lrs = []
        for lr_id, lr_parts in lr_to_parts.items():
            # All have the same type
            lr_type = lr_parts[0]["type"]
            # All have the same x span
            start_x = lr_parts[0]["start_x"]
            stop_x = lr_parts[0]["stop_x"]

            y_values = list(map(lambda part: part["y"], lr_parts))
            # Annotations are 0-index, but openpyxl indexes like excel, starting at 1
            top = min(y_values) + 1
            left = start_x + 1
            bottom = max(y_values) + 1
            right = stop_x + 1

            lrs.append(LabelRegion(lr_id, LabelRegionType(lr_type), top, left, bottom, right))
        return lrs

    def preproces_annotations(self, annotation_file, spreadsheet_file, sheet_name) -> List[LabelRegion]:
        """Reads annotations and returns label regions, which coordinates start at one"""
        logger.debug("Reading Spreadsheet...")
        self._annotation_file = annotation_file
        self._spreadsheet_file = spreadsheet_file
        self._sheet_name = sheet_name
        self._workbook = load_workbook(self._spreadsheet_file)
        self._worksheet = self._workbook[self._sheet_name]
        logger.debug("Reading Annotations...")
        annotation = self._get_annotation()
        logger.debug("Rewriting Chair Label Regions...")
        flattend_lrs = self._flatten_and_rewrite_label_regions(annotation)
        logger.debug("Splitting Chair Label Regions into Cells...")
        cell_rows = self._split_lrs_into_cells(flattend_lrs)

        if self.remove_empty_cells:
            logger.debug("Removing Empty Cells...")
            cell_rows = self._remove_empty_cells(cell_rows)

        logger.debug("Merging Cells into Paper Label Regions...")
        lrs = self._merge_labled_cells_into_lrs(cell_rows)

        return lrs
